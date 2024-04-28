from django.shortcuts import render, redirect
from django.template.loader import render_to_string
from .forms import *
from django.utils import timezone
from django.contrib import messages
from .models import AddStudent, AddexamHall, Examallotment, AddFaculty, AdminAnnounce,AddTimeTable
from django.conf import settings
from django.core.mail import send_mail
from django.shortcuts import get_list_or_404
from itertools import chain
import secrets
import string
import random
from django.db.models import Q
import csv
from django.http import HttpResponse
from django.db import IntegrityError
import pandas as pd
from .forms import ExcelUploadForm
from django.contrib import messages
from .pdffile import *
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment
import re

# To Delete Tables Data in Database
# AddTimeTable.objects.all().delete()
# Examallotment.objects.all().delete()
# AddexamHall.objects.all().delete()
# AddStudent.objects.all().delete()

def addtimetable(request):
    if request.method == 'POST':
        form = AddTimeTableForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('viewtimetable')  
    else:
        form = AddTimeTableForm()
    return render(request, 'addtimetable.html', {'form': form})

def viewtimetable(request):
    timetable_entries = AddTimeTable.objects.all()
    return render(request, 'viewtimetable.html', {'timetable_entries': timetable_entries})





# templates
INDEXPAGE = "index.html"
ADMINLOGINPAGE = "adminlogin.html"
ADMINHOMEPAGE = "adminhome.html"
ADDSTUDENTSPAGE = "addstudents.html"
ADDEXAMHALLSPAGE = "addexamhalls.html"
VIEWSTUDENTSPAGE = "viewstudents.html"
ADDFACULTYPAGE = "addfaculty.html"
VIEWFACULTYPAGE = "viewfaculty.html"
ADDANNOUNCEMENTPAGE = "addannouncement.html"
ADDTIMETABLEPAGE = "addtimetable.html"
VIEWTIMEPABLEPAGE = "viewtimetable.html"
# Create your views here.


def index(req):
    return render(req, INDEXPAGE)


def adminlogin(req):
    context = {}
    context['form'] = AdminlogForm()
    if req.method == "POST":
        form = AdminlogForm(req.POST)
        if form.is_valid():
            adminemail = form.cleaned_data['adminemail']
            adminpassword = form.cleaned_data['adminpassword']
            if adminemail == "admin@gmail.com" and adminpassword == "admin":
                req.session['adminemail'] = adminemail
                return render(req, ADMINHOMEPAGE)
            else:
                messages.warning(req, "Admin Credentials are not Valid......!")
                return render(req, ADMINLOGINPAGE, context)
    return render(req, ADMINLOGINPAGE, context)



def addstudents(req):
    context = {}
    if req.method == "POST":
        form = ExcelUploadForm(req.POST, req.FILES)
        if form.is_valid():
            excel_file = req.FILES['excel_file']
            if excel_file.name.endswith('.xlsx'):
                try:
                    df = pd.read_excel(excel_file)

                    for index, row in df.iterrows():
                        rollnumber = str(row['Roll.No'])
                        name = row['Student Name']
                        department = row['Department']
                        email = row['Email']
                        contact = re.sub(r'\D', '', str(row['Contact']))
                        year = row['Year']
                        semester = row['Semester']
                        profile_url = row['Profile URL']

                        # Generate a random password
                        length = 8
                        characters = string.ascii_letters + string.digits
                        random_password = ''.join(secrets.choice(characters) for _ in range(length))

                        # Create a new instance of the AddStudent model
                        student = AddStudent.objects.create(
                            rollnumber=rollnumber,
                            name=name,
                            department=department,
                            email=email,
                            contact=contact,
                            year=year,
                            semester=semester,
                            profile_url=profile_url,
                            password=random_password
                        )

                        # Sending email with login credentials
                        # subject = "Exam Details"
                        # cont = f'Dear {name},\n\n'
                        # m1 = f"Your Login Credentials:\nUsername: {rollnumber}\nPassword: {random_password}\n\n"
                        # m2 = "Thank you."
                        # m3 = "Regards,\nAdmin."

                        # email_from = settings.EMAIL_HOST_USER
                        # recipient_list = [email]
                        # message = cont + m1 + m2 + m3
                        # send_mail(subject, message, email_from, recipient_list, fail_silently=False)

                    messages.success(req, "Student details added successfully from Excel file.")
                    return redirect('addstudents')

                except Exception as e:
                    messages.error(req, f"Error processing Excel file: {e}")

            else:
                messages.error(req, "Please upload a valid Excel file.")

    else:
        form = ExcelUploadForm()

    context['form'] = form
    return render(req, 'addstudents.html', context)




def addexamhalls(req):
    if req.method == "POST":
        form = AddexamhallForm(req.POST)
        if form.is_valid():
            date = form.cleaned_data['Date']
            starttime = form.cleaned_data['starttime']
            endtime = form.cleaned_data['endtime']
            noofrooms = form.cleaned_data['noofrooms']
            noofbenches = form.cleaned_data['noofbenches']
            students_per_bench = form.cleaned_data.get('students_per_bench')  # Use get method to avoid KeyError
            selected_room_ids = req.POST.getlist('rooms')

            # Validate if students_per_bench is selected
            if not students_per_bench:
                form.add_error('students_per_bench', "Please select students per bench.")
            else:
                rooms = Room.objects.filter(id__in=selected_room_ids)
                selected_rooms = [room.room_number for room in rooms]

                if len(selected_rooms) != noofrooms:
                    form.add_error('rooms', "Number of selected rooms does not match the number entered in the form.")
                else:
                    total_benches = int(noofrooms) * int(noofbenches)  # Ensure integer multiplication
                    total_seats_available = total_benches * int(students_per_bench)  # Ensure integer multiplication
                    print("Total seats available:", total_seats_available)

                    total_students = AddStudent.objects.count()
                    print("Total students:", total_students)

                    if total_seats_available < total_students:
                        messages.error(req, "Not enough seats available in the exam hall for the total number of students.")
                    else:
                        # Create the exam hall instance
                        examhall_data = AddexamHall.objects.create(
                            date=date,
                            starttime=starttime,
                            endtime=endtime,
                            noofrooms=noofrooms,
                            noofbenches=noofbenches,
                            total_benches=total_benches,
                            total_seats=total_seats_available,
                            rooms_list=",".join(selected_rooms),  # Convert list of room numbers to comma-separated string
                            students_per_bench=students_per_bench
                        )
                        messages.success(req, "Exam hall added successfully.")
                        return redirect('addexamhalls')  # Redirect to a success page or another view
        else:
            messages.error(req, "Failed to add exam hall. Please check the form data.")
    else:
        form = AddexamhallForm()
    
    rooms = Room.objects.all()
    
    return render(req, 'addexamhalls.html', {'form': form, 'rooms': rooms})


def delete(req, id):
    print(id)
    AddStudent.objects.filter(id=id).delete()
    return redirect("viewstudents")


def deletefaculty(req,id):
    AddFaculty.objects.filter(id=id).delete()
    return redirect("viewfaculty")



def setseatallotment(request):
    exam_hall_data = AddexamHall.objects.all()
    all_students = AddStudent.objects.order_by('id')  # Order students by their insertion order ID

    exam_halls = []
    for hall in exam_hall_data:
        exam_halls.append({
            'hall': hall,
            'total_benches': hall.total_benches,
            'total_rooms': hall.noofrooms,
            'benches_per_room': hall.noofbenches,
            'students_per_bench': hall.students_per_bench  # Fetch students per bench from the database
        })

    allocated_students = set()

    for hall_info in exam_halls:
        hall = hall_info['hall']
        total_benches = hall_info['total_benches']
        total_rooms = hall_info['total_rooms']
        benches_per_room = hall_info['benches_per_room']
        students_per_bench = hall_info['students_per_bench']

        # Distribute students to benches
        for room_number in range(1, total_rooms + 1):
            allocated_seats = set()  # Track allocated seat numbers in the room
            for bench_number in range(1, benches_per_room + 1):
                branch_count = {}  # Track the count of students from each branch on the bench
                bench_students = []  # Track the students allocated to the current bench
                for seat_number in range(1, students_per_bench + 1):  # Seats per bench
                    student = None
                    # Find a student from a different branch
                    for candidate_student in all_students:
                        if candidate_student.id not in allocated_students and \
                                (branch_count.get(candidate_student.department, 0) < 1):
                            student = candidate_student
                            allocated_students.add(student.id)
                            branch_count[student.department] = branch_count.get(student.department, 0) + 1
                            break  # Allocate only one student per branch
                    if student:
                        # Ensure seat number is unique within the room
                        while seat_number in allocated_seats:
                            seat_number += 1  # Increment by 1
                        # Allocate the student to the seat
                        allocated_seats.add(seat_number)
                        bench_students.append(student)

                # If there are fewer students available than the specified number of students per bench
                if len(bench_students) < students_per_bench:
                    # Fill remaining seats with students from different branches
                    remaining_seats = students_per_bench - len(bench_students)
                    for _ in range(remaining_seats):
                        for candidate_student in all_students:
                            if candidate_student.id not in allocated_students:
                                # Add student from a different branch
                                if candidate_student.department not in branch_count:
                                    student = candidate_student
                                    allocated_students.add(student.id)
                                    branch_count[student.department] = branch_count.get(student.department, 0) + 1
                                    bench_students.append(student)
                                    break

                # Save allocation data to Examallotment model
                for student, seat_number in zip(bench_students, range(1, students_per_bench + 1)):  # Seats per bench
                    room_numbers = hall.rooms_list.split(',')  # Split room numbers from rooms_list field
                    room_display = f"Room{room_numbers[room_number - 1]}"  # Get room number for current room
                    allotment = Examallotment.objects.create(
                        department=student.department,
                        RoomNo=room_display,  # Display room number as "Room112", "Room216", etc.
                        BenchNo=f"Bench{bench_number}",
                        SeatNumber=f"Seat{seat_number}",
                        # Student_Email=student.email,
                        Student_Id=student.rollnumber,
                        date=hall.date,
                        starttime=hall.starttime,
                        endtime=hall.endtime
                    )
                    allotment.save()


                    # Prepare and send email to the student
                    # email_subject = "Exam Seat Allotment Details"
                    # email_message = f"Dear {student.name},\n\nYou have been allotted a seat for your upcoming exam.\n\nBranch: {student.department}\nRoom No: {room_display}\nBench No: {bench_number}\nSeat No: {seat_number}\nDate: {hall.date}\nStart Time: {hall.starttime}\nEnd Time: {hall.endtime}\n\nBest regards,\nAdmin"
                    # send_mail(email_subject, email_message, settings.EMAIL_HOST_USER, [student.email], fail_silently=False)


    return redirect("viewallotedstudents")



def viewallotedstudents(request):
    Exam_alloted_student = Examallotment.objects.all()
    return render(request, "viewallotedstudents.html", {'Exam_alloted_student': Exam_alloted_student})

def viewstudents(req):
    all_students = AddStudent.objects.all()
    return render(req, VIEWSTUDENTSPAGE, {'all_students': all_students})


def addfaculty(req):
    context = {}
    context['form'] = AddFacultyForm()
    if req.method == "POST":
        form = AddFacultyForm(req.POST, req.FILES)
        if form.is_valid():
            length = 8
            characters = string.ascii_letters + string.digits

            # Generate a random password
            random_password = ''.join(secrets.choice(characters) for _ in range(length))
            print("Random Password:", random_password)

            # Extracting form data
            name = form.cleaned_data['name']
            email = form.cleaned_data['email']
            contact = form.cleaned_data['contact']
            branch = form.cleaned_data['branch']
            subject = form.cleaned_data['subject']
            semester = form.cleaned_data['semester']
            year = form.cleaned_data['year']
            image = form.cleaned_data['image']
            profilename = image.name

            try:
                # Attempt to save faculty member
                faculty = AddFaculty.objects.create(
                    name=name,
                    email=email,
                    contact=contact,
                    branch=branch,
                    subject=subject,
                    semester=semester,
                    year=year,
                    image=image,
                    profilename=profilename,
                    password=random_password
                )
            # # Mail Code
            #     # Sending email with login credentials
            #     subject = "Exam Details"
            #     cont = f'Dear {name}'
            #     KEY = f' Branch : {branch}\n'
            #     m1 = f"Your Login Credentials Username : {email}  & password {random_password}"
            #     m2 = "Thanking you"
            #     m3 = "Regards"
            #     m4 = "Admin."

            #     email_from = settings.EMAIL_HOST_USER
            #     recipient_list = [email]
            #     text = cont + '\n' + KEY + '\n' + m1 + '\n' + m2 + '\n' + m3 + '\n' + m4
            #     send_mail(subject, text, email_from, recipient_list, fail_silently=False)

                messages.success(req, "Faculty added successfully")
            except IntegrityError:
                # If email already exists, handle the exception
                messages.warning(req, "A faculty member with the same email already exists")

    return render(req, 'addfaculty.html', context)

def addannouncement(req):
    two_days_content = timezone.now()-timezone.timedelta(days=2)
    messages_to_delete = AdminAnnounce.objects.filter(annuncementdate=two_days_content)
    deleted_data,_ = messages_to_delete.delete()
    all_messages = AdminAnnounce.objects.all()
    context = {}
    context['form'] = AdminAnnouncement()

    if req.method == "POST":
        form = AdminAnnouncement(req.POST)
        print(form.is_valid())
        if form.is_valid():
            announcement = form.cleaned_data['announcement']
            adminemail = req.session['adminemail']
            data = AdminAnnounce(
                announcement=announcement,
                senderemail=adminemail
            )
            data.save()

            # Correct syntax for passing context to the template
            return render(req, ADDANNOUNCEMENTPAGE, {'form': AdminAnnouncement(), 'all_messages': all_messages})

    return render(req, ADDANNOUNCEMENTPAGE, {'form': AdminAnnouncement(), 'all_messages': all_messages})



def viewfaculty(req):
    all_faculty = AddFaculty.objects.all()
    return render(req, VIEWFACULTYPAGE, {'all_faculties': all_faculty})



def download_details(req):
    # Replace YourModel with your actual model
    details_data = Examallotment.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="details.xlsx"'

    # Create a new Excel workbook and add a worksheet
    workbook = Workbook()
    worksheet = workbook.active

    # Add college name, department, and venue information
    worksheet.merge_cells('A1:I1')  # Merge cells for college name
    college_name_cell = worksheet['A1']
    college_name_cell.value = "Aditya Engineering College"
    college_name_cell.font = Font(bold=True, size=18)
    college_name_cell.alignment = Alignment(horizontal='center')


    worksheet.merge_cells('A2:I2')  # Merge cells for venue
    venue_cell = worksheet['A2']
    venue_cell.value = "Venue: BGB"
    venue_cell.font = Font(bold=True)
    venue_cell.alignment = Alignment(horizontal='center')

    # Add header row for exam allotment
    header_row = ['Branch','RoomNo', 'BenchNo', 'SeatNumber',
                   'Student_Id', 'Date', 'Start Time', 'End Time']
    worksheet.append(header_row)

    # Apply bold font to the header row
    for cell in worksheet.iter_rows(min_row=worksheet.max_row, max_row=worksheet.max_row, min_col=1, max_col=len(header_row)):
        for c in cell:
            c.font = Font(bold=True)

    # Write data rows for exam allotment
    for detail in details_data:
        data_row = [detail.department,detail.RoomNo, detail.BenchNo, detail.SeatNumber, 
                    detail.Student_Id, detail.date, detail.starttime, detail.endtime]
        worksheet.append(data_row)

    # Save the workbook to the response
    workbook.save(response)

    return response


